from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook

wb = load_workbook('data.xlsx', data_only=True)

PATH = r"chromedriver.exe"
driver = webdriver.Chrome(PATH)
delay = 3
root_url = 'https://www.olx.co.id/'

for ws in wb.worksheets:
    max_row_a = ws.max_row + 1

    for row_a in range(2, max_row_a):
        url = ws['A' + str(row_a)].value
        url_username = ''
        if url:
            driver.get(url)

            try:
                wait_username = WebDriverWait(driver,
                    delay).until(EC.presence_of_all_elements_located((By.CLASS_NAME, '_3oOe9')))
                url_username = root_url + driver.find_element(By.CLASS_NAME, '_3oOe9').text
            
            except TimeoutException:
                url_username = ''

        ws['B' + str(row_a)].value = url_username

driver.close()
wb.save('hasil_scrap.xlsx') #Simpan sebagai berkas excel baru